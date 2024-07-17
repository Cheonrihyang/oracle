# -*- coding: utf-8 -*-
"""
Created on Wed Jul 10 16:08:39 2024

@author: ORC
"""

#%% json

import json
#json 파일 내용 텍스트 -> 딕셔너리
f=open("user.json","r",encoding=("utf-8"))
dict = json.load(f)
f.close()

#딕셔너리 -> json 파일
user = {"name": "홍길동", "id": 152352, 
    "history": [{"date": "2015-03-11", "item": "iPhone"}, {"date": "2016-02-23", "item": "Monitor"}]
} 

f = open('member.json','w',encoding=('utf-8'))

#json.dump(딕셔너리,파일변수,유니코드 아닌 한글로 인코딩,4칸 들여쓰기)
json.dump(user,f,ensure_ascii=False,indent=4)
f.close()

#%% csv

import csv
#생성
f=open('user.csv','w',encoding=('utf-8'),newline="")
wr=csv.writer(f)
wr.writerow(['ID','이름'])
wr.writerow(['001','뽀로로'])
wr.writerow(['002','크롱'])
f.close()

#조회
f=open('user.csv','r',encoding=('utf-8'))
rdr = csv.reader(f)
next(rdr) #헤더 건너뛰기

for line in rdr:
    print(line)
f.close()

#추가
f=open('user.csv','a',encoding=('utf-8'),newline="")
wr=csv.writer(f)
wr.writerow(['003','에디'])
f.close()

#%% openpyxl

import openpyxl
wb=openpyxl.load_workbook('cha.xlsx')
ws=wb['Sheet1']
ws=wb.active
for row in ws.rows:
    print(row[0].value,row[1].value)
    
ws['B2']='크롱'
ws['B2'].value
wb.save('cha_ver_1.2.xlsx')

#%% xml

import xml.etree.ElementTree as at
#user.xml 문서를 열어서 피싱하여 tree 객체 변수에 대입
tree = at.parse("user.xml")
root = tree.getroot()
lst = root.findall('user')
lst[0].find('name').text

print('사용자수',len(lst))
print('-'*3)
for item in lst:
    print('이름',item.find('name').text)
    print('아이디',item.find('id').text)
    print('수준',item.get('level'))
    print('-'*30)
    
#수정
lst[1].find('name').text='김유신'
lst[1].set('level','3')
#xml 파일저장
tree.write('user2.xml')

#한글문제 해결
header = '<?xml version="1.0" encoding="UTF-8"?>\n'
output = at.tostring(root,encoding="UTF-8")
f=open('user2.xml','w',encoding=("utf-8"))
f.write(header+output.decode('utf-8'))
f.close()

#%% sqllite3

import sqlite3



#테이블 생성
#SQLLite DB 연결
conn = sqlite3.connect('test.db')
#cursor 생성
cur = conn.cursor()
sql='CREATE TABLE IF NOT EXISTS customer (id INTEGER PRIMARY KEY autoincrement , name TEXT,  category INTEGER,region TEXT);'
cur.execute(sql)
conn.close()

#행추가
conn = sqlite3.connect('test.db')
cur = conn.cursor()
sql='INSERT INTO customer(name,category,region) VALUES(:name,:category,:region)'
loc='부산'
#cursor.execute(sq_insert,[리스트],(튜플),{딕셔너리})
cur.execute(sql,('이순신',2,loc))
cur.execute(sql,['강감찬',2,loc])
cur.execute(sql,{'name':'권율','category':3,'region':'인천'})
conn = sqlite3.connect('test.db')


#테이블 변경 후 커밋
conn.commit()
conn.close()

#sql 쿼리 실행
conn = sqlite3.connect('test.db')
cur = conn.cursor()
sql='select * from customer'
cur.execute(sql)
#모든 행(튜플들의 리스트)
rows = cur.fetchall()

#모두 조회
for row in rows:
    print(row)
conn.close()

